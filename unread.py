import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
driver = webdriver.Chrome('C:\ChromeDriver\chromedriver.exe')
driver.get('https://web.whatsapp.com/')

path = r'C:\Users\Sanyam Jain\Downloads\fffggghh.xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb.active

rows = sheet.max_row
col = sheet.max_column

input("Press enter")
def msgtobot(out):
    msg_box = driver.find_element_by_class_name('_3u328')
    msg_box.send_keys(out)
    msg_box.send_keys(Keys.ENTER)
    #msg_box.send_keys("Msg 0 for main menu")
    #msg_box.send_keys(Keys.ENTER)

def replymsg(v):
    reply = "Hi"
    c = 0
    if(str(v)=="1"):
        reply="Hello                                                                                                 " \
              "Please wait while we process your request...\n                                                       " \
              "1.1 : Mechanization \r" \
              "1.2 : Assured prices \r" \
              "1.3 : National Horticulture Mission                                                                   " \
              "1.4 : Pump set                                                                                         "\
              "1.5 : Irrigation system                                                                               " \
              "1.6 : Add on Gadget                                                                                   " \
              "1.7 : Fencing                                                                                         " \
              "1.8 : Assistance on purchase                                                                          " \
              "1.9 : Well                                                                                            " \
              "1.10: Biogas Unit                                                                                     " \
              "1.11: Organic Manure Unit                                                                             " \
              "1.12: Vermicompost Unit                                                                               " \
              "1.13: Water Storage                                                                                   " \
              "1.14: Land Levelling                                                                                  " \
              "1.15: Disilting of Pond                                                                               " \
              "1.16: Development of Jalkund                                                                          " \
              "1.17: Construction of Retaining wall                                                                  " \
              "1.18: Shetkari Adhar Nidhi                                                                            "

    else:
        for i in range(1, rows + 1):
            # for j in range(1, col+1):
            if str(sheet.cell(row=i, column=1).value) == str(v):
                print(sheet.cell(row=i, column=2).value)
                reply = str(sheet.cell(row=i, column=2).value)
            else:
                c = c + 1

    if(c==rows):
        reply = "Type valid input.                                                                                   " \
                "1.Information about schemes                                                                         " \
                "2.Application Status                                                                                " \
                "3.Conatact Info                                                                                     " \
                "4.Krishi card info                                                                                  "


    msgtobot(reply)



    #msg_box = driver.find_element_by_class_name('_3u328')
    #msg_box.send_keys(reply)
    #msg_box.send_keys(Keys.ENTER)



def InputOutput (unread, recipient):
    names = [recipient]
    for name in names:
        #find = name
        person = driver.find_element_by_xpath('//span[@title = "{}"]'.format(name))
        person.click()

        for i in range(1, 4):
            driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
        msg_got = driver.find_elements_by_css_selector("span.selectable-text.invisible-space.copyable-text")
        msg = [message.text for message in msg_got]

        for i in range(-unread, 0):
         print(msg[i])
         replymsg(msg[i])

def loop():
    count = 0
    for i in range(1, 16):
        try:
            c = "//*[@id=\"pane-side\"]/div[1]/div/div/div[" + str(i) + "]/div/div/div[2]/div[2]/div[2]/span[1]/div"
            d = driver.find_element_by_xpath(c)
            x = d.text

            try:
                a = "//*[@id=\"pane-side\"]/div[1]/div/div/div[" + str(i) + "]/div/div/div[2]/div[1]/div[1]/span/span"
                b = driver.find_element_by_xpath(a)
                y = b.text
                print(y)
                print(x)
                InputOutput(int(x), y)
                time.sleep(2)

            except Exception as e1:
               continue
        except Exception as e:
            count = count + 1
            continue
    return count

for a in range(10000000):
    if loop() == 15:
        driver.refresh()
        time.sleep(15)

print("chat bot ended")